#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
쟁승메이드 사업장 회계 장부 자동화 프로그램 v1.0
문의: bgg8988@gmail.com | jaengseung-made.vercel.app
"""

import os
import sys
import json
from datetime import datetime
from typing import Optional

# 필수 패키지 설치 확인
try:
    import pandas as pd
    import openpyxl
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side, numbers
    )
    from openpyxl.utils import get_column_letter
    from openpyxl.styles.numbers import FORMAT_NUMBER_COMMA_SEPARATED1
except ImportError:
    print("필수 패키지가 설치되어 있지 않습니다.")
    print("다음 명령어를 실행하세요: pip install pandas openpyxl")
    sys.exit(1)


# ─────────────────────────────────────────────
# 상수 정의
# ─────────────────────────────────────────────

BRAND_NAME = "쟁승메이드"
BRAND_EMAIL = "bgg8988@gmail.com"
BRAND_URL = "jaengseung-made.vercel.app"
VERSION = "v1.0"

# 업종 정의
INDUSTRIES = {
    "1": "쇼핑몰/이커머스",
    "2": "음식점/카페",
    "3": "제조업",
    "4": "서비스업/프리랜서",
    "5": "기타",
}

# 업종별 수입 카테고리
INCOME_CATEGORIES = {
    "쇼핑몰/이커머스": [
        ("product_sales",    "제품판매"),
        ("shipping_income",  "배송비수입"),
        ("refund_deduct",    "반품환불차감 (마이너스 입력)"),
    ],
    "음식점/카페": [
        ("hall_sales",       "홀매출"),
        ("baemin_sales",     "배달의민족 매출"),
        ("coupang_sales",    "쿠팡이츠 매출"),
        ("yogiyo_sales",     "요기요 매출"),
        ("takeout_sales",    "포장매출"),
    ],
    "제조업": [
        ("product_sales",    "제품판매"),
        ("b2b_sales",        "B2B 납품"),
        ("inventory_change", "원자재 재고변동"),
    ],
    "서비스업/프리랜서": [
        ("project_income",   "프로젝트 수입"),
        ("monthly_income",   "월정액 수입"),
        ("consulting_income","컨설팅 수입"),
    ],
    "기타": [
        ("sales_1",          "매출1"),
        ("sales_2",          "매출2"),
        ("other_income",     "기타 수입"),
    ],
}

# 공통 지출 카테고리
EXPENSE_CATEGORIES = [
    # 매출원가
    ("cogs_goods",       "상품/재료비",          "매출원가",   "변동"),
    ("cogs_outsource",   "외주비",               "매출원가",   "변동"),
    ("cogs_packaging",   "포장비",               "매출원가",   "변동"),
    # 인건비
    ("labor_salary",     "급여",                 "인건비",     "고정"),
    ("labor_insurance",  "4대보험",              "인건비",     "고정"),
    ("labor_severance",  "퇴직금적립",           "인건비",     "고정"),
    # 임대료
    ("rent",             "임대료/관리비",         "임대료",     "고정"),
    # 공과금
    ("util_electricity", "전기요금",             "공과금",     "고정"),
    ("util_water",       "수도요금",             "공과금",     "고정"),
    ("util_gas",         "가스요금",             "공과금",     "고정"),
    ("util_telecom",     "통신비",               "공과금",     "고정"),
    # 마케팅
    ("mkt_ads",          "광고비",               "마케팅",     "변동"),
    ("mkt_platform",     "플랫폼수수료",         "마케팅",     "변동"),
    # 세금
    ("tax_vat",          "부가세",               "세금",       "변동"),
    ("tax_income",       "종합소득세",           "세금",       "변동"),
    ("tax_local",        "지방소득세",           "세금",       "변동"),
    # 기타
    ("etc_supplies",     "소모품비",             "기타",       "변동"),
    ("etc_depreciation", "감가상각비",           "기타",       "고정"),
    ("etc_insurance",    "보험료",               "기타",       "고정"),
]

MONTHS_KR = [
    "1월","2월","3월","4월","5월","6월",
    "7월","8월","9월","10월","11월","12월",
]

QUARTERS_KR = ["1분기 (1~3월)", "2분기 (4~6월)", "3분기 (7~9월)", "4분기 (10~12월)"]


# ─────────────────────────────────────────────
# 유틸리티 함수
# ─────────────────────────────────────────────

def print_banner():
    """프로그램 시작 배너 출력"""
    print("\n" + "=" * 60)
    print(f"  {BRAND_NAME} 사업장 회계 장부 자동화 프로그램 {VERSION}")
    print(f"  문의: {BRAND_EMAIL}")
    print(f"  웹사이트: {BRAND_URL}")
    print("=" * 60 + "\n")


def print_separator(title: str = ""):
    """구분선 출력"""
    if title:
        print(f"\n{'─' * 20} {title} {'─' * 20}")
    else:
        print("─" * 60)


def input_int(prompt: str, default: int = 0) -> int:
    """정수 입력 (기본값 지원)"""
    while True:
        raw = input(prompt).strip()
        if raw == "" and default is not None:
            return default
        try:
            return int(raw.replace(",", ""))
        except ValueError:
            print("  [오류] 숫자만 입력하세요. (천단위 콤마 허용)")


def input_float(prompt: str) -> float:
    """실수 입력"""
    while True:
        raw = input(prompt).strip()
        try:
            return float(raw.replace(",", ""))
        except ValueError:
            print("  [오류] 숫자만 입력하세요.")


def fmt_krw(value: float) -> str:
    """원화 포맷 (천단위 콤마)"""
    if value < 0:
        return f"-{abs(value):,.0f}원"
    return f"{value:,.0f}원"


def fmt_pct(value: float) -> str:
    """퍼센트 포맷"""
    return f"{value:.1f}%"


def select_industry() -> str:
    """업종 선택 메뉴"""
    print_separator("업종 선택")
    for key, name in INDUSTRIES.items():
        print(f"  [{key}] {name}")
    while True:
        choice = input("\n업종을 선택하세요 (1~5): ").strip()
        if choice in INDUSTRIES:
            selected = INDUSTRIES[choice]
            print(f"  → '{selected}' 선택됨\n")
            return selected
        print("  [오류] 1~5 사이의 번호를 입력하세요.")


def select_year() -> int:
    """분석 연도 선택"""
    current_year = datetime.now().year
    year = input_int(f"분석 연도를 입력하세요 (기본값: {current_year}): ", default=current_year)
    return year


# ─────────────────────────────────────────────
# 데이터 수집
# ─────────────────────────────────────────────

def collect_data_manual(industry: str, year: int) -> dict:
    """직접 입력 방식으로 월별 데이터 수집"""
    income_cats = INCOME_CATEGORIES[industry]
    data = {
        "industry": industry,
        "year": year,
        "months": {},
    }

    print_separator(f"{year}년 월별 데이터 입력")
    print("  ※ 금액은 원(₩) 단위로 입력하세요. (천단위 콤마 허용)")
    print("  ※ 입력을 건너뛰려면 그냥 Enter를 누르세요 (0원 처리).\n")

    for month_idx in range(1, 13):
        month_label = MONTHS_KR[month_idx - 1]
        print(f"\n  ── {year}년 {month_label} ──")

        month_data = {"income": {}, "expense": {}}

        # 수입 입력
        print("  [수입]")
        for cat_key, cat_name in income_cats:
            val = input_int(f"    {cat_name}: ", default=0)
            month_data["income"][cat_key] = val

        # 지출 입력
        print("  [지출]")
        for cat_key, cat_name, cat_group, cost_type in EXPENSE_CATEGORIES:
            val = input_int(f"    {cat_name} ({cat_group}): ", default=0)
            month_data["expense"][cat_key] = val

        data["months"][month_idx] = month_data

    return data


def collect_data_excel(industry: str, year: int) -> Optional[dict]:
    """Excel 파일에서 데이터 가져오기"""
    print_separator("Excel 파일에서 가져오기")
    print("  ※ 먼저 메인 메뉴에서 '입력양식 생성'을 선택하여 양식을 만드세요.")
    file_path = input("  가져올 Excel 파일 경로를 입력하세요: ").strip().strip('"')

    if not os.path.exists(file_path):
        print(f"  [오류] 파일을 찾을 수 없습니다: {file_path}")
        return None

    try:
        print("  Excel 파일을 읽는 중...")
        xl = pd.ExcelFile(file_path)
        data = {
            "industry": industry,
            "year": year,
            "months": {},
        }
        income_cats = INCOME_CATEGORIES[industry]

        for month_idx in range(1, 13):
            sheet_name = f"{month_idx}월"
            if sheet_name not in xl.sheet_names:
                # 해당 월 시트가 없으면 0으로 초기화
                month_data = {
                    "income":  {k: 0 for k, _ in income_cats},
                    "expense": {k: 0 for k, _, __, ___ in EXPENSE_CATEGORIES},
                }
                data["months"][month_idx] = month_data
                continue

            df = pd.read_excel(file_path, sheet_name=sheet_name, header=0)
            # 양식: A열=항목키, B열=항목명, C열=금액
            df.columns = ["key", "name", "amount"]
            df["amount"] = pd.to_numeric(df["amount"].fillna(0), errors="coerce").fillna(0)
            amount_map = dict(zip(df["key"].astype(str), df["amount"]))

            month_data = {
                "income":  {k: int(amount_map.get(k, 0)) for k, _ in income_cats},
                "expense": {k: int(amount_map.get(k, 0)) for k, _, __, ___ in EXPENSE_CATEGORIES},
            }
            data["months"][month_idx] = month_data

        print(f"  [완료] {file_path} 파일에서 데이터를 가져왔습니다.")
        return data

    except Exception as e:
        print(f"  [오류] Excel 파일 읽기 실패: {e}")
        return None


# ─────────────────────────────────────────────
# 재무 계산
# ─────────────────────────────────────────────

def calc_month_financials(month_data: dict) -> dict:
    """월별 재무 지표 계산"""
    income = month_data["income"]
    expense = month_data["expense"]

    # 총 매출 (반품은 마이너스로 입력됨)
    total_revenue = sum(income.values())

    # 매출원가
    cogs = (
        expense.get("cogs_goods", 0) +
        expense.get("cogs_outsource", 0) +
        expense.get("cogs_packaging", 0)
    )

    # 매출총이익
    gross_profit = total_revenue - cogs

    # 판매관리비 (인건비 + 임대료 + 공과금 + 마케팅 + 기타)
    sg_and_a = (
        expense.get("labor_salary", 0) +
        expense.get("labor_insurance", 0) +
        expense.get("labor_severance", 0) +
        expense.get("rent", 0) +
        expense.get("util_electricity", 0) +
        expense.get("util_water", 0) +
        expense.get("util_gas", 0) +
        expense.get("util_telecom", 0) +
        expense.get("mkt_ads", 0) +
        expense.get("mkt_platform", 0) +
        expense.get("etc_supplies", 0) +
        expense.get("etc_depreciation", 0) +
        expense.get("etc_insurance", 0)
    )

    # 영업이익
    operating_profit = gross_profit - sg_and_a

    # 세금
    taxes = (
        expense.get("tax_vat", 0) +
        expense.get("tax_income", 0) +
        expense.get("tax_local", 0)
    )

    # 순이익
    net_profit = operating_profit - taxes

    # 총 지출
    total_expense = sum(expense.values())

    # 고정비 / 변동비 분리
    fixed_cost = sum(
        expense.get(k, 0)
        for k, _, __, cost_type in EXPENSE_CATEGORIES
        if cost_type == "고정"
    )
    variable_cost = sum(
        expense.get(k, 0)
        for k, _, __, cost_type in EXPENSE_CATEGORIES
        if cost_type == "변동"
    )

    # 비율 계산 (0 나누기 방지)
    def safe_pct(numerator, denominator):
        if denominator == 0:
            return 0.0
        return round(numerator / denominator * 100, 2)

    gross_margin     = safe_pct(gross_profit, total_revenue)
    operating_margin = safe_pct(operating_profit, total_revenue)
    net_margin       = safe_pct(net_profit, total_revenue)
    labor_ratio      = safe_pct(
        expense.get("labor_salary", 0) + expense.get("labor_insurance", 0) + expense.get("labor_severance", 0),
        total_revenue
    )
    mkt_ratio        = safe_pct(
        expense.get("mkt_ads", 0) + expense.get("mkt_platform", 0),
        total_revenue
    )

    # 손익분기점 (변동비율 기반)
    variable_ratio = safe_pct(variable_cost, total_revenue) / 100
    if variable_ratio < 1:
        bep = round(fixed_cost / (1 - variable_ratio))
    else:
        bep = 0  # 계산 불가

    return {
        "total_revenue":       total_revenue,
        "cogs":                cogs,
        "gross_profit":        gross_profit,
        "sg_and_a":            sg_and_a,
        "operating_profit":    operating_profit,
        "taxes":               taxes,
        "net_profit":          net_profit,
        "total_expense":       total_expense,
        "fixed_cost":          fixed_cost,
        "variable_cost":       variable_cost,
        "gross_margin":        gross_margin,
        "operating_margin":    operating_margin,
        "net_margin":          net_margin,
        "labor_ratio":         labor_ratio,
        "mkt_ratio":           mkt_ratio,
        "bep":                 bep,
    }


def calc_all_financials(data: dict) -> dict:
    """전체 월별 재무 계산 후 분기/연간 집계"""
    results = {}
    for month_idx, month_data in data["months"].items():
        results[month_idx] = calc_month_financials(month_data)

    # 분기별 합산
    quarters = {}
    for q in range(1, 5):
        month_range = range((q - 1) * 3 + 1, q * 3 + 1)
        q_data = {
            key: sum(results[m][key] for m in month_range if m in results)
            for key in results[1].keys()
            if key not in ("gross_margin", "operating_margin", "net_margin", "labor_ratio", "mkt_ratio", "bep")
        }
        # 분기 비율 재계산
        rev = q_data["total_revenue"]
        q_data["gross_margin"]     = round(q_data["gross_profit"] / rev * 100, 2) if rev else 0
        q_data["operating_margin"] = round(q_data["operating_profit"] / rev * 100, 2) if rev else 0
        q_data["net_margin"]       = round(q_data["net_profit"] / rev * 100, 2) if rev else 0
        q_data["labor_ratio"]      = 0
        q_data["mkt_ratio"]        = 0
        q_data["bep"]              = 0
        quarters[q] = q_data

    # 연간 합산
    annual = {
        key: sum(results[m][key] for m in results)
        for key in results[1].keys()
        if key not in ("gross_margin", "operating_margin", "net_margin", "labor_ratio", "mkt_ratio", "bep")
    }
    rev = annual["total_revenue"]
    annual["gross_margin"]     = round(annual["gross_profit"] / rev * 100, 2) if rev else 0
    annual["operating_margin"] = round(annual["operating_profit"] / rev * 100, 2) if rev else 0
    annual["net_margin"]       = round(annual["net_profit"] / rev * 100, 2) if rev else 0
    annual["labor_ratio"]      = 0
    annual["mkt_ratio"]        = 0
    annual["bep"]              = 0

    return {"monthly": results, "quarterly": quarters, "annual": annual}


# ─────────────────────────────────────────────
# 회계 전문가 조언 엔진
# ─────────────────────────────────────────────

def generate_advice(data: dict, financials: dict) -> list:
    """규칙 기반 회계 조언 생성"""
    advice_list = []
    annual = financials["annual"]
    year = data["year"]

    # 1. 마진율 경보
    if annual["gross_margin"] < 20:
        advice_list.append({
            "category": "수익성 경고",
            "priority": "긴급",
            "title": f"매출총이익률 {fmt_pct(annual['gross_margin'])} — 위험 수준",
            "detail": (
                "매출총이익률이 20% 미만입니다. 원가율이 너무 높아 사업 지속성이 위협받을 수 있습니다.\n"
                "• 원재료/상품 공급처 재협상 또는 대체 공급처 탐색\n"
                "• 저마진 제품/서비스 라인 구조조정 검토\n"
                "• 판매가격 인상 가능 여부 시장 조사 필요"
            ),
        })
    elif annual["gross_margin"] < 35:
        advice_list.append({
            "category": "수익성 주의",
            "priority": "주의",
            "title": f"매출총이익률 {fmt_pct(annual['gross_margin'])} — 개선 필요",
            "detail": (
                "매출총이익률이 35% 미만으로 업종 평균 대비 낮은 수준입니다.\n"
                "• 고부가가치 제품/서비스 비중 확대 전략 수립\n"
                "• 원가 절감 가능 항목 면밀히 검토"
            ),
        })

    # 2. 인건비 비율
    labor_total = (
        sum(
            data["months"][m]["expense"].get("labor_salary", 0) +
            data["months"][m]["expense"].get("labor_insurance", 0) +
            data["months"][m]["expense"].get("labor_severance", 0)
            for m in data["months"]
        )
    )
    labor_ratio = labor_total / annual["total_revenue"] * 100 if annual["total_revenue"] else 0
    if labor_ratio > 30:
        advice_list.append({
            "category": "비용 구조",
            "priority": "주의",
            "title": f"인건비 비율 {fmt_pct(labor_ratio)} — 업무 자동화 검토 필요",
            "detail": (
                "인건비가 매출의 30%를 초과합니다. 생산성 향상 방안을 모색하세요.\n"
                "• RPA, 엑셀 자동화 등 업무 자동화 도입으로 인건비 절감\n"
                "• 아웃소싱·프리랜서 활용으로 고정 인건비를 변동비화\n"
                "• 직원별 생산성 KPI 측정 및 성과급 체계 도입"
            ),
        })

    # 3. 마케팅비 비율
    mkt_total = sum(
        data["months"][m]["expense"].get("mkt_ads", 0) +
        data["months"][m]["expense"].get("mkt_platform", 0)
        for m in data["months"]
    )
    mkt_ratio = mkt_total / annual["total_revenue"] * 100 if annual["total_revenue"] else 0
    if mkt_ratio > 10:
        advice_list.append({
            "category": "마케팅 효율",
            "priority": "주의",
            "title": f"마케팅비 비율 {fmt_pct(mkt_ratio)} — ROI 분석 필요",
            "detail": (
                "마케팅비가 매출의 10%를 초과합니다. 채널별 ROI를 정밀 분석하세요.\n"
                "• 채널별 전환율·CAC(고객획득비용) 측정 체계 구축\n"
                "• 효과 낮은 광고 채널 예산 축소, 고ROI 채널에 집중 투자\n"
                "• 리타겟팅·CRM을 통한 재구매율 향상으로 CAC 절감"
            ),
        })

    # 4. 순이익 적자 판단
    loss_months = [m for m, f in financials["monthly"].items() if f["net_profit"] < 0]
    if len(loss_months) >= 6:
        advice_list.append({
            "category": "경영 위기",
            "priority": "긴급",
            "title": f"연간 {len(loss_months)}개월 순손실 — 고정비 구조 즉시 재검토",
            "detail": (
                f"총 {len(loss_months)}개월 순손실이 발생했습니다. 사업 구조 전면 재검토가 필요합니다.\n"
                "• 임대료 협상 또는 이전을 통한 고정비 절감\n"
                "• 손실 유발 서비스/제품 라인 즉시 중단 또는 개편\n"
                "• 단기 자금 확보: 소상공인 정책자금, 신용보증기금 활용 검토"
            ),
        })
    elif annual["net_profit"] < 0:
        advice_list.append({
            "category": "수익성 경고",
            "priority": "긴급",
            "title": f"연간 순손실 {fmt_krw(annual['net_profit'])} — 고정비 구조 재검토",
            "detail": (
                "연간 순손실이 발생했습니다. 고정비 절감이 최우선 과제입니다.\n"
                "• 임대료·인건비 등 고정비 재협상\n"
                "• 불필요한 구독 서비스, 보험료 등 점검"
            ),
        })

    # 5. 부가세 신고 안내
    advice_list.append({
        "category": "세무 일정",
        "priority": "정보",
        "title": "부가세 신고 시기 안내",
        "detail": (
            f"■ {year}년 부가세 신고 일정\n"
            f"  • 1기 확정신고: {year}년 7월 1일 ~ 7월 25일 (1~6월 매출분)\n"
            f"  • 2기 확정신고: {year + 1}년 1월 1일 ~ 1월 25일 (7~12월 매출분)\n\n"
            "■ 준비 서류\n"
            "  • 세금계산서 합계표 (매입/매출)\n"
            "  • 신용카드 매출전표 합계표\n"
            "  • 현금영수증 발행·수취 내역\n\n"
            "  ※ 홈택스(hometax.go.kr)에서 전자신고 가능"
        ),
    })

    # 6. 세금 납부 시뮬레이션 (종합소득세)
    taxable_income = annual["net_profit"]
    if taxable_income > 0:
        # 2024년 기준 종합소득세 세율 구간
        tax_brackets = [
            (14_000_000,  0.06, 0),
            (50_000_000,  0.15, 1_260_000),
            (88_000_000,  0.24, 5_760_000),
            (150_000_000, 0.35, 15_440_000),
            (300_000_000, 0.38, 19_940_000),
            (500_000_000, 0.40, 25_940_000),
            (1_000_000_000, 0.42, 35_940_000),
            (float("inf"),  0.45, 65_940_000),
        ]
        estimated_tax = 0
        for limit, rate, deduction in tax_brackets:
            if taxable_income <= limit:
                estimated_tax = int(taxable_income * rate - deduction)
                break
        local_tax = int(estimated_tax * 0.1)
        advice_list.append({
            "category": "세금 시뮬레이션",
            "priority": "정보",
            "title": f"종합소득세 예상 납부액: {fmt_krw(estimated_tax + local_tax)}",
            "detail": (
                f"■ {year}년 종합소득세 시뮬레이션 (과세표준: {fmt_krw(taxable_income)})\n"
                f"  • 종합소득세: {fmt_krw(estimated_tax)}\n"
                f"  • 지방소득세 (10%): {fmt_krw(local_tax)}\n"
                f"  • 합계 예상 납부액: {fmt_krw(estimated_tax + local_tax)}\n\n"
                "  ※ 각종 공제항목에 따라 실제 납부액은 달라집니다.\n"
                "  ※ 세무사 상담을 통해 절세 전략을 수립하세요."
            ),
        })

    return advice_list


# ─────────────────────────────────────────────
# 콘솔 요약 출력
# ─────────────────────────────────────────────

def print_summary(data: dict, financials: dict, advice_list: list):
    """콘솔에 연간 요약 출력"""
    annual = financials["annual"]
    year = data["year"]

    print_separator(f"{year}년 연간 재무 요약")
    print(f"  업종         : {data['industry']}")
    print(f"  총 매출      : {fmt_krw(annual['total_revenue'])}")
    print(f"  총 지출      : {fmt_krw(annual['total_expense'])}")
    print(f"  매출총이익   : {fmt_krw(annual['gross_profit'])}  ({fmt_pct(annual['gross_margin'])})")
    print(f"  영업이익     : {fmt_krw(annual['operating_profit'])}  ({fmt_pct(annual['operating_margin'])})")
    print(f"  순이익       : {fmt_krw(annual['net_profit'])}  ({fmt_pct(annual['net_margin'])})")

    print_separator("분기별 순이익")
    for q in range(1, 5):
        q_data = financials["quarterly"][q]
        profit = q_data["net_profit"]
        sign = "▲" if profit >= 0 else "▼"
        print(f"  {QUARTERS_KR[q-1]}: {sign} {fmt_krw(profit)}")

    print_separator("회계사 조언 요약")
    for i, adv in enumerate(advice_list, 1):
        print(f"  [{adv['priority']}] {i}. {adv['title']}")


# ─────────────────────────────────────────────
# Excel 스타일 헬퍼
# ─────────────────────────────────────────────

HEADER_FILL = PatternFill(start_color="1D4ED8", end_color="1D4ED8", fill_type="solid")
ALT_FILL    = PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid")
PROFIT_FONT = Font(color="1D4ED8", bold=True)
LOSS_FONT   = Font(color="DC2626", bold=True)
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
BORDER_SIDE = Side(style="thin", color="CBD5E1")
THIN_BORDER = Border(
    left=BORDER_SIDE, right=BORDER_SIDE,
    top=BORDER_SIDE, bottom=BORDER_SIDE,
)
NUM_FMT = '#,##0"원"'
PCT_FMT = '0.0"%"'


def style_header_row(ws, row_num: int, num_cols: int):
    """헤더 행 스타일 적용"""
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER


def style_data_cell(cell, value, is_alt_row: bool = False, is_amount: bool = True, is_pct: bool = False):
    """데이터 셀 스타일 적용"""
    cell.value = value
    cell.border = THIN_BORDER
    cell.alignment = Alignment(horizontal="right" if (is_amount or is_pct) else "left", vertical="center")
    if is_alt_row:
        cell.fill = ALT_FILL
    if is_amount and isinstance(value, (int, float)):
        cell.number_format = NUM_FMT
        if value < 0:
            cell.font = LOSS_FONT
        elif value > 0:
            cell.font = PROFIT_FONT
    elif is_pct and isinstance(value, (int, float)):
        cell.number_format = PCT_FMT


def auto_col_width(ws, min_width: int = 10):
    """열 너비 자동 조정"""
    for col in ws.columns:
        max_len = min_width
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                cell_len = len(str(cell.value)) if cell.value else 0
                if cell_len > max_len:
                    max_len = cell_len
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 30)


# ─────────────────────────────────────────────
# Excel 보고서 생성
# ─────────────────────────────────────────────

def create_report_excel(data: dict, financials: dict, advice_list: list, output_path: str):
    """Excel 보고서 생성 (5개 시트)"""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # 기본 시트 제거

    _sheet1_monthly_pl(wb, data, financials)
    _sheet2_quarterly(wb, data, financials)
    _sheet3_cost_structure(wb, data, financials)
    _sheet4_advice(wb, data, advice_list)
    _sheet5_vat(wb, data, financials)

    wb.save(output_path)
    print(f"\n  [완료] 보고서가 저장되었습니다: {output_path}")


def _sheet1_monthly_pl(wb, data: dict, financials: dict):
    """시트1: 월별 손익계산서"""
    ws = wb.create_sheet("월별 손익계산서")
    year = data["year"]

    # 타이틀
    ws.merge_cells("A1:N1")
    title_cell = ws["A1"]
    title_cell.value = f"{year}년 월별 손익계산서 — {data['industry']}"
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="center")

    # 헤더
    headers = ["항목"] + [f"{m}월" for m in range(1, 13)] + ["연간 합계"]
    for col_idx, h in enumerate(headers, 1):
        ws.cell(row=2, column=col_idx, value=h)
    style_header_row(ws, 2, len(headers))

    # 항목 정의
    rows = [
        ("총 매출",          "total_revenue",    True,  False),
        ("매출원가",         "cogs",             True,  False),
        ("매출총이익",       "gross_profit",     True,  False),
        ("  매출총이익률",   "gross_margin",     False, True),
        ("판매관리비",       "sg_and_a",         True,  False),
        ("영업이익",         "operating_profit", True,  False),
        ("  영업이익률",     "operating_margin", False, True),
        ("세금",             "taxes",            True,  False),
        ("순이익",           "net_profit",       True,  False),
        ("  순이익률",       "net_margin",       False, True),
        ("고정비",           "fixed_cost",       True,  False),
        ("변동비",           "variable_cost",    True,  False),
        ("손익분기점",       "bep",              True,  False),
    ]

    annual = financials["annual"]
    for row_offset, (label, key, is_amount, is_pct) in enumerate(rows):
        row_num = row_offset + 3
        is_alt = row_offset % 2 == 0

        # 항목명
        name_cell = ws.cell(row=row_num, column=1, value=label)
        name_cell.border = THIN_BORDER
        name_cell.alignment = Alignment(horizontal="left")
        if is_alt:
            name_cell.fill = ALT_FILL
        if key in ("gross_profit", "operating_profit", "net_profit"):
            name_cell.font = Font(bold=True)

        # 월별 값
        for m in range(1, 13):
            val = financials["monthly"][m][key]
            cell = ws.cell(row=row_num, column=m + 1)
            style_data_cell(cell, val, is_alt, is_amount, is_pct)

        # 연간 합계
        ann_val = annual[key]
        ann_cell = ws.cell(row=row_num, column=14)
        style_data_cell(ann_cell, ann_val, is_alt, is_amount, is_pct)
        if key in ("gross_profit", "operating_profit", "net_profit"):
            ann_cell.font = Font(
                color="1D4ED8" if ann_val >= 0 else "DC2626",
                bold=True, size=11
            )

    # 틀 고정 (헤더 + 항목명)
    ws.freeze_panes = "B3"
    auto_col_width(ws, min_width=8)
    ws.column_dimensions["A"].width = 16


def _sheet2_quarterly(wb, data: dict, financials: dict):
    """시트2: 분기별 요약"""
    ws = wb.create_sheet("분기별 요약")
    year = data["year"]

    ws.merge_cells("A1:F1")
    ws["A1"].value = f"{year}년 분기별 재무 요약"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")

    headers = ["항목", "1분기 (1~3월)", "2분기 (4~6월)", "3분기 (7~9월)", "4분기 (10~12월)", "연간 합계"]
    for col_idx, h in enumerate(headers, 1):
        ws.cell(row=2, column=col_idx, value=h)
    style_header_row(ws, 2, 6)

    keys = [
        ("총 매출",       "total_revenue",    True,  False),
        ("매출원가",      "cogs",             True,  False),
        ("매출총이익",    "gross_profit",     True,  False),
        ("  이익률",      "gross_margin",     False, True),
        ("영업이익",      "operating_profit", True,  False),
        ("  이익률",      "operating_margin", False, True),
        ("순이익",        "net_profit",       True,  False),
        ("  이익률",      "net_margin",       False, True),
    ]
    annual = financials["annual"]
    for row_offset, (label, key, is_amount, is_pct) in enumerate(keys):
        row_num = row_offset + 3
        is_alt = row_offset % 2 == 0

        name_cell = ws.cell(row=row_num, column=1, value=label)
        name_cell.border = THIN_BORDER
        name_cell.alignment = Alignment(horizontal="left")
        if is_alt:
            name_cell.fill = ALT_FILL

        for q in range(1, 5):
            val = financials["quarterly"][q][key]
            cell = ws.cell(row=row_num, column=q + 1)
            style_data_cell(cell, val, is_alt, is_amount, is_pct)

        ann_cell = ws.cell(row=row_num, column=6)
        style_data_cell(ann_cell, annual[key], is_alt, is_amount, is_pct)

    ws.freeze_panes = "B3"
    auto_col_width(ws, min_width=12)
    ws.column_dimensions["A"].width = 14


def _sheet3_cost_structure(wb, data: dict, financials: dict):
    """시트3: 비용 구조 분석"""
    ws = wb.create_sheet("비용 구조 분석")
    year = data["year"]
    annual_revenue = financials["annual"]["total_revenue"]

    ws.merge_cells("A1:E1")
    ws["A1"].value = f"{year}년 비용 구조 분석 — 항목별 금액 및 매출 비율"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")

    headers = ["비용 항목", "분류", "성격", "연간 금액", "매출 비율"]
    for col_idx, h in enumerate(headers, 1):
        ws.cell(row=2, column=col_idx, value=h)
    style_header_row(ws, 2, 5)

    # 항목별 연간 합산
    annual_expenses = {}
    for m, month_data in data["months"].items():
        for k, v in month_data["expense"].items():
            annual_expenses[k] = annual_expenses.get(k, 0) + v

    for row_offset, (cat_key, cat_name, cat_group, cost_type) in enumerate(EXPENSE_CATEGORIES):
        row_num = row_offset + 3
        is_alt = row_offset % 2 == 0
        amount = annual_expenses.get(cat_key, 0)
        ratio = (amount / annual_revenue * 100) if annual_revenue else 0

        for col_idx, val in enumerate([cat_name, cat_group, cost_type, amount, ratio], 1):
            cell = ws.cell(row=row_num, column=col_idx)
            is_amount = col_idx == 4
            is_pct = col_idx == 5
            style_data_cell(cell, val, is_alt, is_amount, is_pct)
            if col_idx <= 3:
                cell.alignment = Alignment(horizontal="left")

    # 합계 행
    total_row = len(EXPENSE_CATEGORIES) + 3
    total_expense = sum(annual_expenses.values())
    total_ratio = (total_expense / annual_revenue * 100) if annual_revenue else 0

    ws.cell(row=total_row, column=1, value="합  계").font = Font(bold=True)
    total_amt_cell = ws.cell(row=total_row, column=4, value=total_expense)
    total_amt_cell.number_format = NUM_FMT
    total_amt_cell.font = Font(bold=True)
    total_pct_cell = ws.cell(row=total_row, column=5, value=total_ratio)
    total_pct_cell.number_format = PCT_FMT
    total_pct_cell.font = Font(bold=True)

    for col in range(1, 6):
        ws.cell(row=total_row, column=col).border = THIN_BORDER
        ws.cell(row=total_row, column=col).fill = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")

    auto_col_width(ws, min_width=10)
    ws.column_dimensions["A"].width = 16
    ws.freeze_panes = "A3"


def _sheet4_advice(wb, data: dict, advice_list: list):
    """시트4: 회계사 조언 리포트"""
    ws = wb.create_sheet("회계사 조언 리포트")
    year = data["year"]

    ws.merge_cells("A1:D1")
    ws["A1"].value = f"{year}년 회계 전문가 조언 리포트 — {data['industry']}"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:D2")
    ws["A2"].value = f"생성일시: {datetime.now().strftime('%Y-%m-%d %H:%M')}  |  {BRAND_NAME} ({BRAND_URL})"
    ws["A2"].font = Font(size=10, color="64748B")
    ws["A2"].alignment = Alignment(horizontal="center")

    headers = ["No.", "분류", "우선순위", "조언 제목 및 상세 내용"]
    for col_idx, h in enumerate(headers, 1):
        ws.cell(row=3, column=col_idx, value=h)
    style_header_row(ws, 3, 4)

    priority_colors = {
        "긴급": "FEE2E2",
        "주의": "FEF3C7",
        "정보": "EFF6FF",
    }
    priority_font_colors = {
        "긴급": "DC2626",
        "주의": "D97706",
        "정보": "1D4ED8",
    }

    current_row = 4
    for i, adv in enumerate(advice_list, 1):
        # 상세 내용은 줄 수에 따라 행 병합
        detail_lines = adv["detail"].split("\n")
        row_height = max(len(detail_lines) * 15, 40)

        for col_idx in range(1, 5):
            cell = ws.cell(row=current_row, column=col_idx)
            cell.border = THIN_BORDER
            fill_color = priority_colors.get(adv["priority"], "FFFFFF")
            cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

        ws.cell(row=current_row, column=1, value=i).alignment = Alignment(horizontal="center", vertical="top")
        ws.cell(row=current_row, column=2, value=adv["category"]).alignment = Alignment(horizontal="center", vertical="top")

        pri_cell = ws.cell(row=current_row, column=3, value=adv["priority"])
        pri_cell.alignment = Alignment(horizontal="center", vertical="top")
        pri_cell.font = Font(
            color=priority_font_colors.get(adv["priority"], "000000"),
            bold=True
        )

        content = f"■ {adv['title']}\n\n{adv['detail']}"
        content_cell = ws.cell(row=current_row, column=4, value=content)
        content_cell.alignment = Alignment(
            horizontal="left", vertical="top", wrap_text=True
        )

        ws.row_dimensions[current_row].height = row_height
        current_row += 1

    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 70
    ws.freeze_panes = "A4"


def _sheet5_vat(wb, data: dict, financials: dict):
    """시트5: 부가세 신고 준비 자료"""
    ws = wb.create_sheet("부가세 신고 준비")
    year = data["year"]

    ws.merge_cells("A1:F1")
    ws["A1"].value = f"{year}년 부가세 신고 준비 자료"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")

    # 1기 (1~6월) / 2기 (7~12월) 분리
    periods = {
        f"1기 (1~6월) — {year}년 7월 신고": range(1, 7),
        f"2기 (7~12월) — {year + 1}년 1월 신고": range(7, 13),
    }

    income_cats = INCOME_CATEGORIES[data["industry"]]
    current_row = 3

    for period_label, month_range in periods.items():
        # 기간 헤더
        ws.merge_cells(f"A{current_row}:F{current_row}")
        period_cell = ws.cell(row=current_row, column=1, value=period_label)
        period_cell.fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
        period_cell.font = Font(color="FFFFFF", bold=True, size=12)
        period_cell.alignment = Alignment(horizontal="center")
        current_row += 1

        # 소계 헤더
        headers = ["구분", "항목", "과세표준(공급가액)", "세율", "부가세액", "비고"]
        for col_idx, h in enumerate(headers, 1):
            ws.cell(row=current_row, column=col_idx, value=h)
        style_header_row(ws, current_row, 6)
        current_row += 1

        # 매출 내역
        period_revenue = 0
        for cat_key, cat_name in income_cats:
            period_amount = sum(
                data["months"][m]["income"].get(cat_key, 0)
                for m in month_range if m in data["months"]
            )
            supply_value = int(period_amount / 1.1)  # 공급가액 (부가세 포함 금액 기준)
            vat = period_amount - supply_value
            period_revenue += period_amount

            is_alt = (current_row % 2 == 0)
            for col_idx in range(1, 7):
                cell = ws.cell(row=current_row, column=col_idx)
                cell.border = THIN_BORDER
                if is_alt:
                    cell.fill = ALT_FILL

            ws.cell(row=current_row, column=1, value="매출").alignment = Alignment(horizontal="center")
            ws.cell(row=current_row, column=2, value=cat_name)
            supply_cell = ws.cell(row=current_row, column=3, value=supply_value)
            supply_cell.number_format = NUM_FMT
            ws.cell(row=current_row, column=4, value=0.10).number_format = PCT_FMT
            vat_cell = ws.cell(row=current_row, column=5, value=vat)
            vat_cell.number_format = NUM_FMT
            ws.cell(row=current_row, column=6, value="과세")
            current_row += 1

        # 매입 세금계산서 관련 비용
        vatable_expenses = [
            ("cogs_goods",    "상품/재료비"),
            ("cogs_outsource","외주비"),
            ("mkt_ads",       "광고비"),
            ("etc_supplies",  "소모품비"),
        ]
        for exp_key, exp_name in vatable_expenses:
            period_amount = sum(
                data["months"][m]["expense"].get(exp_key, 0)
                for m in month_range if m in data["months"]
            )
            if period_amount == 0:
                continue
            supply_value = int(period_amount / 1.1)
            vat = period_amount - supply_value

            is_alt = (current_row % 2 == 0)
            for col_idx in range(1, 7):
                cell = ws.cell(row=current_row, column=col_idx)
                cell.border = THIN_BORDER
                if is_alt:
                    cell.fill = ALT_FILL

            ws.cell(row=current_row, column=1, value="매입").alignment = Alignment(horizontal="center")
            ws.cell(row=current_row, column=2, value=exp_name)
            supply_cell = ws.cell(row=current_row, column=3, value=supply_value)
            supply_cell.number_format = NUM_FMT
            ws.cell(row=current_row, column=4, value=0.10).number_format = PCT_FMT
            vat_cell = ws.cell(row=current_row, column=5, value=vat)
            vat_cell.number_format = NUM_FMT
            ws.cell(row=current_row, column=6, value="매입공제")
            current_row += 1

        current_row += 2  # 간격

    auto_col_width(ws, min_width=10)
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["F"].width = 12


# ─────────────────────────────────────────────
# 입력 양식 Excel 생성
# ─────────────────────────────────────────────

def create_input_template(industry: str, year: int, output_path: str):
    """업종별 입력 양식 Excel 생성"""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    income_cats = INCOME_CATEGORIES[industry]

    for month_idx in range(1, 13):
        sheet_name = f"{month_idx}월"
        ws = wb.create_sheet(sheet_name)

        # 타이틀
        ws.merge_cells("A1:C1")
        ws["A1"].value = f"{year}년 {MONTHS_KR[month_idx - 1]} 입력 양식 — {industry}"
        ws["A1"].font = Font(bold=True, size=12)
        ws["A1"].alignment = Alignment(horizontal="center")

        # 컬럼 헤더
        for col_idx, h in enumerate(["항목 키", "항목명", "금액 (원)"], 1):
            ws.cell(row=2, column=col_idx, value=h)
        style_header_row(ws, 2, 3)

        # 수입 섹션
        ws.cell(row=3, column=1, value="──── 수입 ────")
        ws.cell(row=3, column=1).font = Font(bold=True, color="1D4ED8")
        ws.merge_cells(f"A3:C3")
        ws["A3"].fill = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
        ws["A3"].alignment = Alignment(horizontal="center")

        row = 4
        for cat_key, cat_name in income_cats:
            ws.cell(row=row, column=1, value=cat_key)
            ws.cell(row=row, column=2, value=cat_name)
            amount_cell = ws.cell(row=row, column=3, value=0)
            amount_cell.number_format = NUM_FMT
            for col in range(1, 4):
                ws.cell(row=row, column=col).border = THIN_BORDER
            row += 1

        # 지출 섹션
        ws.cell(row=row, column=1, value="──── 지출 ────")
        ws.merge_cells(f"A{row}:C{row}")
        ws[f"A{row}"].font = Font(bold=True, color="DC2626")
        ws[f"A{row}"].fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
        ws[f"A{row}"].alignment = Alignment(horizontal="center")
        row += 1

        for cat_key, cat_name, cat_group, cost_type in EXPENSE_CATEGORIES:
            ws.cell(row=row, column=1, value=cat_key)
            ws.cell(row=row, column=2, value=f"[{cat_group}] {cat_name} ({cost_type})")
            amount_cell = ws.cell(row=row, column=3, value=0)
            amount_cell.number_format = NUM_FMT
            for col in range(1, 4):
                ws.cell(row=row, column=col).border = THIN_BORDER
            row += 1

        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 30
        ws.column_dimensions["C"].width = 18
        ws.freeze_panes = "A3"

    wb.save(output_path)
    print(f"  [완료] 입력 양식이 생성되었습니다: {output_path}")
    print("  ※ C열 (금액) 셀에 숫자를 입력하고 저장하세요.")


# ─────────────────────────────────────────────
# 메인 실행 흐름
# ─────────────────────────────────────────────

def main():
    print_banner()

    # 작업 디렉토리 설정
    work_dir = os.getcwd()
    print(f"  작업 디렉토리: {work_dir}\n")

    # ── 메인 메뉴 ──
    while True:
        print_separator("메인 메뉴")
        print("  [1] 새 회계 분석 시작")
        print("  [2] 입력 양식(Excel) 생성")
        print("  [0] 종료")
        choice = input("\n선택하세요: ").strip()

        if choice == "0":
            print("\n  프로그램을 종료합니다. 감사합니다!\n")
            break

        elif choice == "2":
            # 입력 양식 생성
            industry = select_industry()
            year = select_year()
            template_path = os.path.join(
                work_dir,
                f"입력양식_{industry}_{year}.xlsx"
            )
            create_input_template(industry, year, template_path)

        elif choice == "1":
            # 분석 시작
            industry = select_industry()
            year = select_year()

            # 데이터 입력 방식 선택
            print_separator("데이터 입력 방식")
            print("  [A] 직접 입력 (프로그램에서 월별 금액 입력)")
            print("  [B] Excel 파일에서 가져오기")
            input_method = input("\n선택하세요 (A/B): ").strip().upper()

            data = None
            if input_method == "B":
                data = collect_data_excel(industry, year)
            if data is None:  # A 선택 or B 실패 시 직접 입력
                if input_method == "B":
                    print("  → 직접 입력 방식으로 전환합니다.")
                data = collect_data_manual(industry, year)

            # 재무 계산
            print("\n  재무 지표를 계산하는 중...")
            financials = calc_all_financials(data)

            # 조언 생성
            advice_list = generate_advice(data, financials)

            # 콘솔 요약 출력
            print_summary(data, financials, advice_list)

            # Excel 보고서 출력
            timestamp = datetime.now().strftime("%Y%m%d_%H%M")
            report_path = os.path.join(
                work_dir,
                f"회계보고서_{industry}_{year}_{timestamp}.xlsx"
            )

            print_separator("Excel 보고서 생성")
            print(f"  보고서 저장 경로: {report_path}")
            create_report_excel(data, financials, advice_list, report_path)

            # 완료 메시지
            print_separator("완료")
            print(f"  보고서가 성공적으로 생성되었습니다!")
            print(f"  파일: {report_path}")
            print("\n  포함된 시트:")
            print("    • 시트1: 월별 손익계산서 (12개월)")
            print("    • 시트2: 분기별 요약")
            print("    • 시트3: 비용 구조 분석")
            print("    • 시트4: 회계사 조언 리포트")
            print("    • 시트5: 부가세 신고 준비 자료")

            # 데이터 저장 여부 (JSON 백업)
            save_json = input("\n  입력 데이터를 JSON으로 백업하시겠습니까? (y/N): ").strip().lower()
            if save_json == "y":
                json_path = report_path.replace(".xlsx", "_data.json")
                with open(json_path, "w", encoding="utf-8") as f:
                    json.dump(data, f, ensure_ascii=False, indent=2)
                print(f"  [완료] 데이터 백업: {json_path}")

        else:
            print("  [오류] 0, 1, 2 중 하나를 입력하세요.")


# ─────────────────────────────────────────────
# 진입점
# ─────────────────────────────────────────────

if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\n\n  [중단] 사용자가 프로그램을 종료했습니다.")
        print("  저장된 데이터는 유지됩니다.\n")
        sys.exit(0)
    except Exception as e:
        print(f"\n  [오류] 예기치 않은 오류가 발생했습니다: {e}")
        print("  문의: bgg8988@gmail.com\n")
        sys.exit(1)


# ═══════════════════════════════════════════════════════════════
#  쟁승메이드 (JaengseungMade) — 프리미엄 개발 서비스
#  ▸ 웹사이트 : jaengseung-made.vercel.app
#  ▸ 이메일   : bgg8988@gmail.com
#  ▸ 연락처   : 010-3907-1392
#  7년차 대기업 백엔드 개발자가 만드는 실전 자동화 솔루션
# ═══════════════════════════════════════════════════════════════
