#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
================================================================================
  웹 크롤링 자동화 도구 v1.0
  Made by 쟁승메이드  |  jaengseung-made.com
================================================================================

  필요 패키지 설치:
    pip install requests beautifulsoup4 openpyxl lxml

  사용법:
    1. 아래 ── 설정 ── 영역에서 TARGET_URL과 옵션을 수정하세요.
    2. 터미널에서 실행:  python web_scraper_v1.0.py
    3. 같은 폴더에 엑셀 결과 파일이 저장됩니다.

  지원 기능:
    - 단일/다중 페이지 크롤링 (페이지네이션 자동 탐색)
    - 재시도 로직 (네트워크 오류 자동 재시도)
    - 엑셀 저장 (서식 포함)
    - 요청 간격 조절 (서버 부하 방지)
    - 로그 출력 및 파일 저장

  맞춤 개발이 필요하다면: jaengseung-made.com/freelance
================================================================================
"""

import requests
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
import time
import logging
import sys
import os

# ── 설정 (이 부분을 수정하세요) ───────────────────────────────────────────────

TARGET_URL     = "https://example.com"       # 크롤링할 URL
DELAY_SECONDS  = 1.5                          # 요청 간격 (서버 부하 방지, 최소 1.0 권장)
MAX_PAGES      = 5                            # 최대 크롤링 페이지 수 (1 = 단일 페이지)
OUTPUT_FILE    = f"크롤링결과_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
LOG_TO_FILE    = True                         # True: 로그 파일도 저장

# 페이지네이션 설정 (다음 페이지 링크 선택자 — 사이트마다 다름)
NEXT_PAGE_SELECTOR = "a.next, a[rel='next'], .pagination .next a"

# 데이터 추출 설정 (아래 extract_data 함수에서 상세 수정)
# 기본: 페이지 내 모든 링크 수집 (예시용)

# ────────────────────────────────────────────────────────────────────────────

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/120.0.0.0 Safari/537.36"
    ),
    "Accept":          "text/html,application/xhtml+xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "ko-KR,ko;q=0.9,en;q=0.8",
    "Accept-Encoding": "gzip, deflate, br",
}


def setup_logger() -> logging.Logger:
    handlers: list[logging.Handler] = [logging.StreamHandler(sys.stdout)]
    if LOG_TO_FILE:
        log_name = f"scraper_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log"
        handlers.append(logging.FileHandler(log_name, encoding="utf-8"))
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s [%(levelname)s] %(message)s",
        handlers=handlers,
    )
    return logging.getLogger(__name__)


logger = setup_logger()


def fetch_page(url: str, retries: int = 3) -> BeautifulSoup | None:
    """페이지 가져오기 (재시도 포함)"""
    for attempt in range(retries):
        try:
            resp = requests.get(url, headers=HEADERS, timeout=15)
            resp.raise_for_status()
            resp.encoding = resp.apparent_encoding or "utf-8"
            logger.info(f"✅ 페이지 로드 성공: {url}")
            return BeautifulSoup(resp.text, "lxml")
        except requests.exceptions.HTTPError as e:
            logger.warning(f"HTTP 오류 [{attempt+1}/{retries}]: {e}")
        except requests.exceptions.ConnectionError:
            logger.warning(f"연결 오류 [{attempt+1}/{retries}]: {url}")
        except requests.exceptions.Timeout:
            logger.warning(f"시간 초과 [{attempt+1}/{retries}]: {url}")
        except Exception as e:
            logger.warning(f"알 수 없는 오류 [{attempt+1}/{retries}]: {e}")

        if attempt < retries - 1:
            wait = DELAY_SECONDS * (attempt + 2)
            logger.info(f"  → {wait:.1f}초 후 재시도...")
            time.sleep(wait)

    logger.error(f"❌ 페이지 로드 실패: {url}")
    return None


def extract_data(soup: BeautifulSoup, page_url: str) -> list[dict]:
    """
    ========================================================================
    페이지에서 데이터를 추출합니다.

    🔧 이 함수를 목적에 맞게 수정하세요!

    예시 1 — 제품 목록 수집:
        for item in soup.select(".product-item"):
            name  = item.select_one(".product-name")
            price = item.select_one(".product-price")
            items.append({
                "상품명":  name.get_text(strip=True) if name else "",
                "가격":    price.get_text(strip=True) if price else "",
                "수집URL": page_url,
            })

    예시 2 — 뉴스 기사 수집:
        for article in soup.select("article, .news-item"):
            title = article.select_one("h2, h3, .title")
            date  = article.select_one(".date, time")
            items.append({
                "제목": title.get_text(strip=True) if title else "",
                "날짜": date.get_text(strip=True) if date else "",
            })
    ========================================================================
    """
    items = []
    collected_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # ── 기본 예시: 모든 링크 텍스트 수집 ──────────────────────
    for link in soup.find_all("a", href=True):
        text = link.get_text(strip=True)
        href = link["href"]
        # 너무 짧거나 빈 텍스트 제외
        if text and len(text) >= 3:
            items.append({
                "링크 텍스트": text[:200],
                "URL":        href[:500],
                "출처 페이지": page_url,
                "수집 시간":  collected_at,
            })

    return items


def get_next_page_url(soup: BeautifulSoup, current_url: str) -> str | None:
    """다음 페이지 URL 반환 (없으면 None)"""
    next_link = soup.select_one(NEXT_PAGE_SELECTOR)
    if not next_link:
        return None

    href = next_link.get("href", "")
    if not href:
        return None

    # 상대 URL → 절대 URL 변환
    if href.startswith("http"):
        return href
    from urllib.parse import urljoin
    return urljoin(current_url, href)


def save_to_excel(data: list[dict], filepath: str) -> None:
    """엑셀 저장 (서식 포함)"""
    if not data:
        logger.warning("⚠️ 저장할 데이터가 없습니다.")
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "크롤링 결과"

    # 헤더 스타일
    header_fill  = PatternFill(start_color="1D4ED8", end_color="1D4ED8", fill_type="solid")
    header_font  = Font(color="FFFFFF", bold=True, size=11, name="맑은 고딕")
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border  = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    headers = list(data[0].keys())
    ws.row_dimensions[1].height = 28

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill      = header_fill
        cell.font      = header_font
        cell.alignment = center_align
        cell.border    = thin_border

    # 데이터 입력
    row_font = Font(size=10, name="맑은 고딕")
    for row_idx, item in enumerate(data, start=2):
        for col_idx, key in enumerate(headers, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=item.get(key, ""))
            cell.font      = row_font
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border    = thin_border
        # 짝수 행 배경색
        if row_idx % 2 == 0:
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = PatternFill(
                    start_color="EFF6FF", end_color="EFF6FF", fill_type="solid"
                )

    # 열 너비 자동 조정
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)

    # 첫 행 고정
    ws.freeze_panes = "A2"

    # 요약 시트
    ws_summary = wb.create_sheet("요약")
    ws_summary["A1"] = "크롤링 요약"
    ws_summary["A1"].font = Font(bold=True, size=14)
    ws_summary["A3"] = "수집 URL"
    ws_summary["B3"] = TARGET_URL
    ws_summary["A4"] = "수집 건수"
    ws_summary["B4"] = len(data)
    ws_summary["A5"] = "수집 일시"
    ws_summary["B5"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws_summary["A7"] = "Made by 쟁승메이드 | jaengseung-made.com"
    ws_summary["A7"].font = Font(color="6B7280", size=9)

    wb.save(filepath)
    logger.info(f"✅ 저장 완료: {filepath}  ({len(data):,}건)")


def main():
    logger.info("=" * 60)
    logger.info("  웹 크롤링 자동화 도구 v1.0  |  쟁승메이드")
    logger.info("=" * 60)
    logger.info(f"대상 URL: {TARGET_URL}")
    logger.info(f"최대 페이지: {MAX_PAGES}")

    all_data: list[dict] = []
    current_url: str | None = TARGET_URL

    for page_num in range(1, MAX_PAGES + 1):
        if not current_url:
            break

        logger.info(f"\n[페이지 {page_num}/{MAX_PAGES}] {current_url}")
        soup = fetch_page(current_url)
        if not soup:
            break

        page_data = extract_data(soup, current_url)
        all_data.extend(page_data)
        logger.info(f"  → 수집: {len(page_data)}건 (누적: {len(all_data)}건)")

        current_url = get_next_page_url(soup, current_url)
        if current_url and page_num < MAX_PAGES:
            logger.info(f"  → 다음 페이지 대기 {DELAY_SECONDS}초...")
            time.sleep(DELAY_SECONDS)

    logger.info(f"\n총 수집: {len(all_data):,}건")
    save_to_excel(all_data, OUTPUT_FILE)

    logger.info("\n완료! 맞춤 자동화 개발이 필요하다면: jaengseung-made.com")


if __name__ == "__main__":
    main()
