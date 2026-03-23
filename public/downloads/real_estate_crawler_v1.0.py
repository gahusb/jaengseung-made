#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
===========================================================
  부동산 매물 크롤링 프로그램 v1.0
  쟁승메이드 | jaengseung-made.vercel.app
  문의: bgg8988@gmail.com | 010-3907-1392
===========================================================
  지원 플랫폼:
    - 직방 (Zigbang)
    - 다방 (Dabang)
    - 피터팬 (Peterpanz)
    - 네이버 부동산 (Naver Land)

  수집 결과: Excel 파일 (.xlsx) — 플랫폼별 시트 + 전체 시트

  필요 패키지 설치:
    pip install requests beautifulsoup4 pandas openpyxl tqdm
===========================================================
"""

import sys
import time
import json
import re
import random
from datetime import datetime
from pathlib import Path

# ── 패키지 확인 ────────────────────────────────────────────
try:
    import requests
    from bs4 import BeautifulSoup
    import pandas as pd
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError as e:
    print(f"\n[오류] 필요한 패키지가 없습니다: {e}")
    print("아래 명령어로 설치 후 다시 실행하세요:")
    print("  pip install requests beautifulsoup4 pandas openpyxl tqdm\n")
    sys.exit(1)


# ══════════════════════════════════════════════════════════
#  공통 설정
# ══════════════════════════════════════════════════════════
HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/122.0.0.0 Safari/537.36"
    ),
    "Accept": "application/json, text/plain, */*",
    "Accept-Language": "ko-KR,ko;q=0.9,en-US;q=0.8",
    "Accept-Encoding": "gzip, deflate, br",
    "Connection": "keep-alive",
}

DELAY_MIN = 1.2
DELAY_MAX = 2.8


def _delay():
    """요청 간 랜덤 딜레이 — 서버 부하 방지 및 차단 회피"""
    time.sleep(random.uniform(DELAY_MIN, DELAY_MAX))


# ══════════════════════════════════════════════════════════
#  데이터 스키마
# ══════════════════════════════════════════════════════════
COLUMNS = [
    "플랫폼", "매물유형", "거래유형",
    "보증금(만원)", "월세(만원)", "가격표시",
    "면적(㎡)", "층수", "주소", "건물명",
    "설명", "등록일", "링크",
]


def _item(**kwargs) -> dict:
    base = {c: "" for c in COLUMNS}
    base.update(kwargs)
    return base


def _fmt(amount) -> str:
    """정수(만원) → '3억 2,000만' or '1,500만' 문자열"""
    try:
        amount = int(amount)
    except (TypeError, ValueError):
        return str(amount)
    if amount <= 0:
        return "0"
    if amount >= 10000:
        eok = amount // 10000
        man = amount % 10000
        return f"{eok}억" if man == 0 else f"{eok}억 {man:,}만"
    return f"{amount:,}만"


def _price_str(deposit, rent, trade_type: str) -> str:
    if trade_type == "월세":
        return f"{_fmt(deposit)} / {_fmt(rent)}"
    if trade_type == "전세":
        return f"전세 {_fmt(deposit)}"
    return _fmt(deposit)


# ══════════════════════════════════════════════════════════
#  위경도 → Geohash 변환 (직방 API용)
# ══════════════════════════════════════════════════════════
def _to_geohash(lat: float, lng: float, precision: int = 5) -> str:
    BASE32 = "0123456789bcdefghjkmnpqrstuvwxyz"
    lat_r, lng_r = [-90.0, 90.0], [-180.0, 180.0]
    gh, bits, bit, even = [], 0, 0, True
    while len(gh) < precision:
        if even:
            mid = (lng_r[0] + lng_r[1]) / 2
            if lng >= mid:
                bits = (bits << 1) | 1; lng_r[0] = mid
            else:
                bits <<= 1; lng_r[1] = mid
        else:
            mid = (lat_r[0] + lat_r[1]) / 2
            if lat >= mid:
                bits = (bits << 1) | 1; lat_r[0] = mid
            else:
                bits <<= 1; lat_r[1] = mid
        even = not even
        bit += 1
        if bit == 5:
            gh.append(BASE32[bits]); bits = bit = 0
    return "".join(gh)


# ══════════════════════════════════════════════════════════
#  주요 지역 위경도 DB
# ══════════════════════════════════════════════════════════
LOCATION_DB: dict[str, tuple[float, float]] = {
    # 서울 자치구
    "강남구": (37.5172, 127.0473), "서초구": (37.4837, 127.0324),
    "송파구": (37.5145, 127.1059), "마포구": (37.5638, 126.9084),
    "강서구": (37.5509, 126.8495), "영등포구": (37.5264, 126.8962),
    "용산구": (37.5311, 126.9810), "성동구": (37.5634, 127.0369),
    "동작구": (37.5124, 126.9393), "관악구": (37.4784, 126.9516),
    "강북구": (37.6396, 127.0255), "노원구": (37.6543, 127.0568),
    "은평구": (37.6026, 126.9291), "서대문구": (37.5791, 126.9368),
    "종로구": (37.5735, 126.9789), "중구": (37.5641, 126.9979),
    "성북구": (37.5894, 127.0167), "광진구": (37.5385, 127.0823),
    "중랑구": (37.5953, 127.0843), "도봉구": (37.6688, 127.0471),
    "강동구": (37.5301, 127.1238), "구로구": (37.4954, 126.8874),
    "금천구": (37.4600, 126.9000), "양천구": (37.5170, 126.8666),
    "동대문구": (37.5744, 127.0396),
    # 주요 동/지역
    "강남": (37.5172, 127.0473), "홍대": (37.5563, 126.9233),
    "신촌": (37.5596, 126.9361), "이태원": (37.5346, 126.9944),
    "합정": (37.5495, 126.9140), "연남동": (37.5607, 126.9224),
    "성수동": (37.5443, 127.0557), "건대": (37.5402, 127.0702),
    "혜화": (37.5820, 127.0019), "신림": (37.4843, 126.9292),
    "사당": (37.4762, 126.9815), "잠실": (37.5133, 127.1000),
    "여의도": (37.5216, 126.9245), "명동": (37.5607, 126.9862),
    # 수도권
    "인천": (37.4563, 126.7052), "수원": (37.2636, 127.0286),
    "성남": (37.4201, 127.1267), "분당": (37.3825, 127.1175),
    "판교": (37.3943, 127.1106), "일산": (37.6577, 126.7670),
    "부천": (37.5035, 126.7660), "안양": (37.3943, 126.9568),
    "의정부": (37.7381, 127.0337), "평택": (36.9921, 127.1130),
    # 지방 광역시
    "부산": (35.1796, 129.0756), "대구": (35.8714, 128.6014),
    "대전": (36.3504, 127.3845), "광주": (35.1595, 126.8526),
    "울산": (35.5384, 129.3114), "세종": (36.4800, 127.2890),
}

# 네이버 부동산 법정동 코드 주요 목록
CORTAR_CODES: dict[str, str] = {
    "서울 강남구":      "1168000000",
    "서울 서초구":      "1165000000",
    "서울 송파구":      "1171000000",
    "서울 마포구":      "1144000000",
    "서울 용산구":      "1117000000",
    "서울 영등포구":    "1156000000",
    "서울 강서구":      "1150000000",
    "서울 관악구":      "1162000000",
    "서울 성동구":      "1120000000",
    "서울 동작구":      "1159000000",
    "서울 노원구":      "1135000000",
    "서울 은평구":      "1138000000",
    "서울 강동구":      "1174000000",
    "경기 성남시 분당구": "4113500000",
    "경기 수원시 팔달구": "4111700000",
    "경기 고양시 일산동구": "4128100000",
    "부산 해운대구":    "2635000000",
    "부산 수영구":      "2644000000",
    "대구 수성구":      "2723000000",
    "인천 연수구":      "2817000000",
}


def get_coords(address: str) -> tuple[float, float] | None:
    for key, coords in LOCATION_DB.items():
        if key in address:
            return coords
    return None


# ══════════════════════════════════════════════════════════
#  1. 직방 크롤러
# ══════════════════════════════════════════════════════════
class ZigbangCrawler:
    """직방 내부 API 기반 매물 수집"""

    API = "https://apis.zigbang.com"

    ROOM_CODE = {
        "원룸": "원룸", "투룸": "투쓰리룸", "오피스텔": "오피스텔",
        "아파트": "아파트", "빌라": "빌라",
    }
    TRADE_CODE = {"월세": "월세", "전세": "전세", "매매": "매매"}

    def _geocode(self, address: str) -> tuple[float, float] | None:
        """직방 자체 검색 API로 위경도 조회"""
        try:
            r = requests.get(
                f"{self.API}/v2/suggest",
                params={"q": address, "serviceType": "all"},
                headers=HEADERS, timeout=10,
            )
            for item in r.json().get("items", []):
                lat = item.get("lat") or item.get("latitude")
                lng = item.get("lng") or item.get("longitude")
                if lat and lng:
                    return float(lat), float(lng)
        except Exception:
            pass
        return None

    def fetch(self, address: str, room_type: str, trade_type: str, limit: int = 100) -> list[dict]:
        tag = f"[직방] {room_type}/{trade_type}"
        print(f"  {tag} 수집 중...", end=" ", flush=True)

        coords = get_coords(address) or self._geocode(address)
        if not coords:
            print(f"위치 찾기 실패")
            return []

        lat, lng = coords
        gh = _to_geohash(lat, lng, 5)
        stype = self.ROOM_CODE.get(room_type, "원룸")
        ttype = self.TRADE_CODE.get(trade_type, "월세")

        url = (
            f"{self.API}/v2/items"
            f"?deposit_gteq=0&rent_gteq=0"
            f"&sales_type_in={ttype}"
            f"&service_type_eq={stype}"
            f"&geohash={gh}&domain=zigbang&withCoords=true&per_page={limit}"
        )

        results = []
        try:
            _delay()
            r = requests.get(url, headers={**HEADERS, "Referer": "https://www.zigbang.com/"}, timeout=15)
            for it in r.json().get("items", []):
                dep = int(it.get("deposit") or 0)
                rent = int(it.get("rent") or 0)
                item_id = it.get("item_id", "")
                results.append(_item(
                    플랫폼="직방",
                    매물유형=room_type,
                    거래유형=trade_type,
                    **{"보증금(만원)": dep, "월세(만원)": rent},
                    가격표시=_price_str(dep, rent, trade_type),
                    **{"면적(㎡)": round(float(it.get("area") or 0), 1) or ""},
                    층수=str(it.get("floor") or ""),
                    주소=it.get("address") or it.get("local1") or "",
                    건물명=it.get("title") or "",
                    링크=f"https://www.zigbang.com/home/oneroom/items/{item_id}" if item_id else "",
                ))
        except Exception as e:
            print(f"오류({e})")
            return results

        print(f"{len(results)}건")
        return results


# ══════════════════════════════════════════════════════════
#  2. 다방 크롤러
# ══════════════════════════════════════════════════════════
class DabangCrawler:
    """다방 내부 API 기반 매물 수집"""

    API = "https://www.dabangapp.com/api/3"

    ROOM_CODE = {"원룸": 1, "투룸": 2, "쓰리룸": 3, "오피스텔": 10, "아파트": 11, "빌라": 21}
    TRADE_CODE = {"월세": 1, "전세": 2, "매매": 3}

    def fetch(self, lat: float, lng: float, room_type: str, trade_type: str, limit: int = 80) -> list[dict]:
        tag = f"[다방] {room_type}/{trade_type}"
        print(f"  {tag} 수집 중...", end=" ", flush=True)

        rc = self.ROOM_CODE.get(room_type, 1)
        tc = self.TRADE_CODE.get(trade_type, 1)
        delta = 0.025

        params = {
            "call_type": "web",
            "device_type": "web",
            "filters": json.dumps({
                "sellingTypeList": [tc],
                "roomTypeList": [rc],
                "depositRange": {"min": 0, "max": 90000},
                "priceRange": {"min": 0, "max": 999},
                "areaRange": {"min": 0, "max": 999},
            }),
            "location": json.dumps([
                [lat - delta, lng - delta],
                [lat + delta, lng + delta],
            ]),
            "page": 1,
            "limit": limit,
            "ordering": "-created_at",
        }

        results = []
        try:
            _delay()
            r = requests.get(
                f"{self.API}/room/list/map", params=params,
                headers={**HEADERS, "Referer": "https://www.dabangapp.com/"}, timeout=15,
            )
            data = r.json()
            rooms = data.get("rooms") or data.get("result", {}).get("rooms") or []
            for room in rooms:
                price = room.get("price") or [0, 0]
                dep = int(price[0]) if isinstance(price, list) and price else 0
                rent = int(price[1]) if isinstance(price, list) and len(price) > 1 else 0
                rid = room.get("id", "")
                results.append(_item(
                    플랫폼="다방",
                    매물유형=room_type,
                    거래유형=trade_type,
                    **{"보증금(만원)": dep, "월세(만원)": rent},
                    가격표시=_price_str(dep, rent, trade_type),
                    **{"면적(㎡)": round(float(room.get("area") or 0), 1) or ""},
                    층수=str(room.get("floor") or ""),
                    주소=room.get("address") or "",
                    건물명=room.get("name") or "",
                    설명=(room.get("title") or "")[:80],
                    링크=f"https://www.dabangapp.com/room/{rid}" if rid else "",
                ))
        except Exception as e:
            print(f"오류({e})")
            return results

        print(f"{len(results)}건")
        return results


# ══════════════════════════════════════════════════════════
#  3. 피터팬 크롤러
# ══════════════════════════════════════════════════════════
class PeterpanzCrawler:
    """피터팬의 좋은방구하기 크롤러 (API + HTML 폴백)"""

    BASE = "https://www.peterpanz.com"
    TRADE_CODE = {"월세": "monthly", "전세": "charter", "매매": "selling"}

    def fetch(self, lat: float, lng: float, trade_type: str, max_pages: int = 3) -> list[dict]:
        tag = f"[피터팬] {trade_type}"
        print(f"  {tag} 수집 중...", end=" ", flush=True)

        tc = self.TRADE_CODE.get(trade_type, "monthly")
        results = []

        for page in range(1, max_pages + 1):
            try:
                _delay()
                r = requests.get(
                    f"{self.BASE}/house/all",
                    params={"lat": lat, "lng": lng, "zoom": 13, "type": tc, "page": page},
                    headers={**HEADERS, "Referer": f"{self.BASE}/"},
                    timeout=15,
                )

                # JSON 응답 시도
                try:
                    data = r.json()
                    houses = data.get("houses") or data.get("items") or []
                    if not houses:
                        break
                    for h in houses:
                        hid = h.get("idx") or h.get("id") or ""
                        dep = int(h.get("deposit") or 0)
                        rent = int(h.get("rent") or 0)
                        results.append(_item(
                            플랫폼="피터팬",
                            매물유형=h.get("type") or "",
                            거래유형=trade_type,
                            **{"보증금(만원)": dep, "월세(만원)": rent},
                            가격표시=_price_str(dep, rent, trade_type),
                            **{"면적(㎡)": h.get("area") or ""},
                            층수=str(h.get("floor") or ""),
                            주소=h.get("address") or "",
                            건물명=h.get("title") or "",
                            설명=(h.get("description") or "")[:80],
                            링크=f"{self.BASE}/house/detail/{hid}" if hid else "",
                        ))
                except (json.JSONDecodeError, ValueError):
                    # HTML 파싱 폴백
                    soup = BeautifulSoup(r.text, "html.parser")
                    cards = soup.select(".item-list li, .house-list-item, [class*='item']")
                    if not cards:
                        break
                    for card in cards:
                        title_el = card.select_one(".title, .name, h3, h4")
                        price_el = card.select_one(".price, .cost, [class*='price']")
                        addr_el = card.select_one(".address, .addr, [class*='address']")
                        link_el = card.select_one("a[href]")
                        href = link_el["href"] if link_el else ""
                        results.append(_item(
                            플랫폼="피터팬",
                            거래유형=trade_type,
                            가격표시=price_el.get_text(strip=True) if price_el else "",
                            주소=addr_el.get_text(strip=True) if addr_el else "",
                            건물명=title_el.get_text(strip=True) if title_el else "",
                            링크=self.BASE + href if href.startswith("/") else href,
                        ))

            except Exception as e:
                print(f"페이지{page} 오류({e}) ", end="")
                break

        print(f"{len(results)}건")
        return results


# ══════════════════════════════════════════════════════════
#  4. 네이버 부동산 크롤러
# ══════════════════════════════════════════════════════════
class NaverLandCrawler:
    """네이버 부동산 내부 API 기반 매물 수집 (법정동 코드 필요)"""

    API = "https://new.land.naver.com/api"

    TRADE_CODE = {"매매": "A1", "전세": "B1", "월세": "B2,B3"}
    TYPE_CODE = {
        "아파트": "APT", "오피스텔": "OPST",
        "빌라": "VL", "원룸": "OR", "상가": "SG",
    }

    def fetch(self, cortar_no: str, real_estate_type: str, trade_type: str, max_pages: int = 5) -> list[dict]:
        tag = f"[네이버부동산] {real_estate_type}/{trade_type}"
        print(f"  {tag} 수집 중...", end=" ", flush=True)

        tc = self.TRADE_CODE.get(trade_type, "A1")
        rc = self.TYPE_CODE.get(real_estate_type, "APT")

        naver_headers = {
            **HEADERS,
            "Referer": "https://new.land.naver.com/",
            "Accept": "*/*",
        }

        results = []
        for page in range(1, max_pages + 1):
            try:
                _delay()
                r = requests.get(
                    f"{self.API}/articles",
                    params={
                        "cortarNo": cortar_no,
                        "order": "rank",
                        "realEstateType": rc,
                        "tradeType": tc,
                        "page": page,
                        "pageSize": 20,
                    },
                    headers=naver_headers,
                    timeout=15,
                )

                if r.status_code != 200:
                    if r.status_code == 403:
                        print(f"\n  [네이버부동산] 봇 차단 감지 (403). 잠시 후 재시도하거나 VPN 사용 권장.")
                    break

                data = r.json()
                articles = data.get("articleList") or []
                if not articles:
                    break

                for art in articles:
                    ano = art.get("articleNo") or ""
                    dep_str = art.get("dealOrWarrantPrc") or art.get("warrantyPrice") or ""
                    rent_str = art.get("rentPrc") or ""

                    results.append(_item(
                        플랫폼="네이버부동산",
                        매물유형=art.get("realEstateTypeName") or real_estate_type,
                        거래유형=trade_type,
                        가격표시=dep_str + (f" / {rent_str}만" if rent_str else ""),
                        **{"면적(㎡)": art.get("area1") or ""},
                        층수=art.get("floorInfo") or "",
                        주소=art.get("cortarAddress") or "",
                        건물명=art.get("articleName") or "",
                        설명=(art.get("articleFeatureDesc") or "")[:80],
                        등록일=art.get("articleConfirmYmd") or "",
                        링크=f"https://new.land.naver.com/houses?articleNo={ano}" if ano else "",
                    ))

            except Exception as e:
                print(f"페이지{page} 오류({e}) ", end="")
                break

        print(f"{len(results)}건")
        return results


# ══════════════════════════════════════════════════════════
#  Excel 내보내기
# ══════════════════════════════════════════════════════════
COL_WIDTHS = {
    "플랫폼": 10, "매물유형": 10, "거래유형": 8,
    "보증금(만원)": 14, "월세(만원)": 12, "가격표시": 22,
    "면적(㎡)": 10, "층수": 8, "주소": 32, "건물명": 22,
    "설명": 38, "등록일": 12, "링크": 55,
}

HEADER_FILL = PatternFill("solid", fgColor="1D4ED8")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
ALT_FILL    = PatternFill("solid", fgColor="EFF6FF")
CELL_BORDER = Border(
    left=Side(style="thin", color="CBD5E1"),
    right=Side(style="thin", color="CBD5E1"),
    top=Side(style="thin", color="CBD5E1"),
    bottom=Side(style="thin", color="CBD5E1"),
)


def _style_sheet(ws) -> None:
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = CELL_BORDER

    for i, row in enumerate(ws.iter_rows(min_row=2), 2):
        for cell in row:
            if i % 2 == 0:
                cell.fill = ALT_FILL
            cell.border = CELL_BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=False)

    for col_idx, col_name in enumerate(COLUMNS, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = COL_WIDTHS.get(col_name, 15)

    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def export_excel(items: list[dict], filename: str) -> None:
    if not items:
        print("\n[경고] 수집된 데이터가 없습니다.")
        return

    df = pd.DataFrame(items, columns=COLUMNS)

    with pd.ExcelWriter(filename, engine="openpyxl") as writer:
        # 전체 시트
        df.to_excel(writer, sheet_name="전체", index=False)

        # 플랫폼별 시트
        for platform in df["플랫폼"].unique():
            sub = df[df["플랫폼"] == platform]
            sub.to_excel(writer, sheet_name=platform[:31], index=False)

        # 요약 시트
        summary = (
            df.groupby(["플랫폼", "거래유형", "매물유형"])
            .size().reset_index(name="건수")
            .sort_values("건수", ascending=False)
        )
        summary.to_excel(writer, sheet_name="요약", index=False)

        # 스타일 적용
        for ws in writer.book.worksheets:
            _style_sheet(ws)

    total = len(df)
    platforms = df["플랫폼"].unique().tolist()
    print(f"\n✅ 저장 완료: {filename}")
    print(f"   총 {total}건 | 플랫폼: {', '.join(platforms)}")


# ══════════════════════════════════════════════════════════
#  메인 인터페이스
# ══════════════════════════════════════════════════════════
BANNER = """
╔══════════════════════════════════════════════════╗
║    부동산 매물 크롤링 프로그램  v1.0             ║
║    쟁승메이드 · jaengseung-made.vercel.app       ║
╚══════════════════════════════════════════════════╝
"""


def _choose(prompt: str, options: list[str]) -> str:
    print(f"\n{prompt}")
    for i, o in enumerate(options, 1):
        print(f"  {i}. {o}")
    while True:
        try:
            n = int(input("▶ 선택: ").strip())
            if 1 <= n <= len(options):
                return options[n - 1]
        except (ValueError, KeyboardInterrupt):
            pass
        print("  올바른 번호를 입력하세요.")


def _choose_multi(prompt: str, options: list[str]) -> list[str]:
    print(f"\n{prompt} (콤마 구분, 예: 1,3 / 전체: 0)")
    for i, o in enumerate(options, 1):
        print(f"  {i}. {o}")
    raw = input("▶ 선택: ").strip()
    if raw == "0" or not raw:
        return list(options)
    chosen = []
    for part in raw.split(","):
        try:
            n = int(part.strip())
            if 1 <= n <= len(options):
                chosen.append(options[n - 1])
        except ValueError:
            pass
    return chosen if chosen else list(options)


def main() -> None:
    print(BANNER)

    # 1. 지역 입력
    print("📍 검색할 지역을 입력하세요")
    print("   예: 강남구 / 마포구 / 홍대 / 수원 / 부산")
    print("   (DB에 없는 지역은 위경도 직접 입력 가능)")
    address = input("▶ 지역: ").strip() or "강남구"

    coords = get_coords(address)
    if not coords:
        print(f"\n  '{address}'이(가) 기본 DB에 없습니다.")
        try:
            lat = float(input("  위도(lat) 입력 (예: 37.5172): ").strip())
            lng = float(input("  경도(lng) 입력 (예: 127.0473): ").strip())
            coords = (lat, lng)
        except ValueError:
            print("  좌표 입력 실패. 강남구 기본값으로 진행합니다.")
            coords = (37.5172, 127.0473)

    lat, lng = coords
    print(f"  → 위치: {lat:.4f}, {lng:.4f}")

    # 2. 거래 유형
    trade_type = _choose("💰 거래 유형", ["월세", "전세", "매매"])

    # 3. 매물 유형
    room_types = _choose_multi(
        "🏠 매물 유형 (복수 선택 가능)",
        ["원룸", "투룸", "오피스텔", "아파트", "빌라"],
    )

    # 4. 플랫폼 선택
    platforms = _choose_multi(
        "🌐 수집 플랫폼",
        ["직방", "다방", "피터팬", "네이버부동산"],
    )

    # 네이버 부동산 법정동 코드 입력
    cortar_no = ""
    naver_types: list[str] = []
    if "네이버부동산" in platforms:
        print("\n  [네이버부동산] 법정동 코드가 필요합니다.")
        print("  알려진 코드 목록:")
        for name, code in list(CORTAR_CODES.items())[:8]:
            print(f"    {name:20s} : {code}")
        cortar_no = input("  법정동 코드 입력 (없으면 Enter 건너뜀): ").strip()
        if cortar_no:
            naver_types = [rt for rt in room_types if rt in NaverLandCrawler.TYPE_CODE]
            if not naver_types:
                print("  선택한 매물 유형 중 네이버부동산 지원 유형 없음 — 건너뜀")
                cortar_no = ""

    # ── 수집 시작 ──────────────────────────────────────────
    print(f"\n{'─'*50}")
    print(f"🔍 수집 시작")
    print(f"   지역: {address}  |  거래: {trade_type}")
    print(f"   매물: {', '.join(room_types)}")
    print(f"   플랫폼: {', '.join(platforms)}")
    print(f"{'─'*50}\n")

    all_items: list[dict] = []

    if "직방" in platforms:
        zc = ZigbangCrawler()
        for rt in room_types:
            all_items.extend(zc.fetch(address, rt, trade_type))

    if "다방" in platforms:
        dc = DabangCrawler()
        for rt in room_types:
            all_items.extend(dc.fetch(lat, lng, rt, trade_type))

    if "피터팬" in platforms:
        pc = PeterpanzCrawler()
        all_items.extend(pc.fetch(lat, lng, trade_type, max_pages=3))

    if "네이버부동산" in platforms and cortar_no:
        nc = NaverLandCrawler()
        for rt in naver_types:
            all_items.extend(nc.fetch(cortar_no, rt, trade_type))

    # 중복 제거 (링크 기준)
    seen: set[str] = set()
    deduped: list[dict] = []
    for item in all_items:
        key = item.get("링크") or f"{item.get('주소')}{item.get('건물명')}{item.get('가격표시')}"
        if key and key not in seen:
            seen.add(key)
            deduped.append(item)

    print(f"\n{'─'*50}")
    print(f"📊 수집 완료: {len(all_items)}건 → 중복 제거 후 {len(deduped)}건")

    if not deduped:
        print("  수집된 매물이 없습니다. 지역·플랫폼 옵션을 변경해보세요.")
        return

    # 저장
    timestamp = datetime.now().strftime("%Y%m%d_%H%M")
    safe_addr = re.sub(r'[\\/:*?"<>|]', "_", address)
    filename = f"부동산매물_{safe_addr}_{trade_type}_{timestamp}.xlsx"

    print(f"💾 저장 중: {filename}")
    export_excel(deduped, filename)

    print(f"\n{'═'*50}")
    print("  프로그램 완료!")
    print(f"  문의: bgg8988@gmail.com  |  010-3907-1392")
    print(f"  쟁승메이드: jaengseung-made.vercel.app")
    print(f"{'═'*50}\n")


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\n\n사용자가 프로그램을 종료했습니다.")
